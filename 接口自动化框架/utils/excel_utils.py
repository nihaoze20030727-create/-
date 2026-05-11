import openpyxl

from config.config import *


def read_excel(file_path=EXCEL_FILE,sheet_name=SHEET_NAME):

    # workbook = openpyxl.load_workbook("./data/测试用例.xlsx")
    workbook = openpyxl.load_workbook(file_path)

    worksheet = workbook[sheet_name]

    data = []
    keys = [cell.value for cell in worksheet[2]]
    for row in worksheet.iter_rows(min_row=3,values_only=True):
        dict_data = dict(zip(keys, row))
        if dict_data["is_true"]:
            data.append(dict_data)

    workbook.close()

    return data

# read_excel()

